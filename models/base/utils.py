# %%
import torch
import torch.nn as nn

from collections import OrderedDict
import numpy as np
import os
from models.base.parameter import Parameter

# %%
def summary(model, input_size, batch_size=-1, device=torch.device('cuda:0'), dtypes=None):
    result, params_info = summary_string(
        model, input_size, batch_size, device, dtypes)
    print(result)

    return params_info


def summary_string(model, input_size, batch_size=-1, device=torch.device('cuda:0'), dtypes=None):
    if dtypes == None:
        dtypes = [torch.FloatTensor]*len(input_size)

    summary_str = ''

    def register_hook(module):
        def hook(module, input, output):
            class_name = str(module.__class__).split(".")[-1].split("'")[0]
            module_idx = len(summary)

            m_key = "%s-%i" % (class_name, module_idx + 1)
            summary[m_key] = OrderedDict()
            summary[m_key]["input_shape"] = list(input[0].size())
            summary[m_key]["input_shape"][0] = batch_size
            if isinstance(output, (list, tuple)):
                summary[m_key]["output_shape"] = [
                    [-1] + list(o.size())[1:] for o in output
                ]
            else:
                summary[m_key]["output_shape"] = list(output.size())
                summary[m_key]["output_shape"][0] = batch_size

            params = 0
            if hasattr(module, "weight") and hasattr(module.weight, "size"):
                params += torch.prod(torch.LongTensor(list(module.weight.size())))
                summary[m_key]["trainable"] = module.weight.requires_grad
            if hasattr(module, "bias") and hasattr(module.bias, "size"):
                params += torch.prod(torch.LongTensor(list(module.bias.size())))
            summary[m_key]["nb_params"] = params

        if (
            not isinstance(module, nn.Sequential)
            and not isinstance(module, nn.ModuleList)
        ):
            hooks.append(module.register_forward_hook(hook))

    # multiple inputs to the network
    if isinstance(input_size, tuple):
        input_size = [input_size]

    # batch_size of 2 for batchnorm
    x = [torch.rand(2, *in_size).type(dtype).to(device=device)
         for in_size, dtype in zip(input_size, dtypes)]

    # create properties
    summary = OrderedDict()
    hooks = []

    # register hook
    model.apply(register_hook)

    # make a forward pass
    # print(x.shape)
    model(*x)

    # remove these hooks
    for h in hooks:
        h.remove()

    summary_str += "----------------------------------------------------------------" + "\n"
    line_new = "{:>20}  {:>25} {:>15}".format(
        "Layer (type)", "Output Shape", "Param #")
    summary_str += line_new + "\n"
    summary_str += "================================================================" + "\n"
    total_params = 0
    total_output = 0
    trainable_params = 0
    for layer in summary:
        # input_shape, output_shape, trainable, nb_params
        line_new = "{:>20}  {:>25} {:>15}".format(
            layer,
            str(summary[layer]["output_shape"]),
            "{0:,}".format(summary[layer]["nb_params"]),
        )
        total_params += summary[layer]["nb_params"]

        total_output += np.prod(summary[layer]["output_shape"])
        if "trainable" in summary[layer]:
            if summary[layer]["trainable"] == True:
                trainable_params += summary[layer]["nb_params"]
        summary_str += line_new + "\n"

    # assume 4 bytes/number (float on cuda).
    total_input_size = abs(np.prod(sum(input_size, ()))
                           * batch_size * 4. / (1024 ** 2.))
    total_output_size = abs(2. * total_output * 4. /
                            (1024 ** 2.))  # x2 for gradients
    total_params_size = abs(total_params * 4. / (1024 ** 2.))
    total_size = total_params_size + total_output_size + total_input_size

    summary_str += "================================================================" + "\n"
    summary_str += "Total params: {0:,}".format(total_params) + "\n"
    summary_str += "Trainable params: {0:,}".format(trainable_params) + "\n"
    summary_str += "Non-trainable params: {0:,}".format(total_params -
                                                        trainable_params) + "\n"
    summary_str += "----------------------------------------------------------------" + "\n"
    summary_str += "Input size (MB): %0.2f" % total_input_size + "\n"
    summary_str += "Forward/backward pass size (MB): %0.2f" % total_output_size + "\n"
    summary_str += "Params size (MB): %0.2f" % total_params_size + "\n"
    summary_str += "Estimated Total Size (MB): %0.2f" % total_size + "\n"
    summary_str += "----------------------------------------------------------------" + "\n"
    # return summary
    return summary_str, (total_params, trainable_params)

# %%
class EarlyStopping:
    """Early stops the training if validation loss doesn't improve after a given patience."""
    def __init__(self, save_path, patience=5, verbose=False, delta=0):
        """
        Args:
            patience (int): How long to wait after last time validation loss improved.
                            Default: 5
            verbose (bool): If True, prints a message for each validation loss improvement. 
                            Default: False
            delta (float): Minimum change in the monitored quantity to qualify as an improvement.
                            Default: 0
        """
        self.patience = patience
        self.verbose = verbose
        self.counter = 0
        self.best_score = None
        self.early_stop = False
        self.val_loss_min = np.Inf
        self.delta = delta
        self.save_path = save_path

    def __call__(self, val_loss, model):

        score = -val_loss

        if self.best_score is None:
            self.best_score = score
            self.save_checkpoint(val_loss, model)
        elif score < self.best_score + self.delta:
            self.counter += 1
            print(f'EarlyStopping counter: {self.counter} out of {self.patience}')
            if self.counter >= self.patience:
                self.early_stop = True
        else:
            self.best_score = score
            self.save_checkpoint(val_loss, model)
            self.counter = 0

    def save_checkpoint(self, val_loss, model):
        '''Saves model when validation loss decrease.'''
        if self.verbose:
            print(f'Validation loss decreased ({self.val_loss_min:.6f} --> {val_loss:.6f}).  Saving model ...')
        path = os.path.join(self.save_path, 'best_network.pth')
        torch.save(model, path)	# 这里会存储迄今最优模型的参数
        self.val_loss_min = val_loss


# %%
def onehot(indices: torch.Tensor, num_classes: int) -> torch.Tensor:
    """Convert a tensor of indices of any shape `(N, ...)` to a
    tensor of one-hot indicators of shape `(N, num_classes, ...)` and of type uint8. Output's device is equal to the
    input's device`.
    Args:
        indices: input tensor to convert.
        num_classes: number of classes for one-hot tensor.
    .. versionchanged:: 0.4.3
        This functions is now torchscriptable.
    """
    new_shape = (indices.shape[0], num_classes) + indices.shape[1:]
    onehot = torch.zeros(new_shape, dtype=torch.uint8, device=indices.device)
    return onehot.scatter_(1, indices.unsqueeze(1), 1)


# %%
class Metrics:
    def __init__(self, pred:torch.Tensor, true:torch.Tensor, n_class):
        pred = pred.reshape(-1, n_class)
        true = true.flatten()
        # print(pred.shape)
        # print(true.shape)
        assert pred.shape[0] == true.shape[0]
        assert pred.shape[1] == n_class
        pred, true = pred.detach(), true.detach()

        # target is (batch_size, ...)
        pred = torch.argmax(pred, dim=1).flatten()
        true = true.flatten()

        mask = (true >= 0) & (true < n_class)
        self.true = true[mask]
        self.pred = pred[mask]
        self.n_class = n_class

    
    def confusion_matrix(self):
        if not hasattr(self, 'cm'):
            indices = self.n_class * self.true + self.pred
            self.cm = torch.bincount(indices, minlength=self.n_class ** 2).reshape(self.n_class, self.n_class)
            self.cm = self.cm.float()
        return self.cm


    def accuracy_score(self):
        if not hasattr(self, 'acc'):
            count = torch.sum(torch.eq(self.true, self.pred), dtype=torch.float32)
            self.acc = count/self.true.shape[0]
        return self.acc


    def precision_score(self):
        if not hasattr(self, 'pr'):
            self.pr = torch.sum(self.confusion_matrix(), dim=0)
            for i in range(self.n_class):
                self.pr[i] = self.confusion_matrix()[i, i]/self.pr[i]
        return self.pr


    def recall_score(self):
        if not hasattr(self, 're'):
            self.re = torch.sum(self.confusion_matrix(), dim=1)
            for i in range(self.n_class):
                self.re[i] = self.confusion_matrix()[i, i]/self.re[i]
        return self.re


    def f1_score(self):
        if not hasattr(self, 'f1'):
            self.f1 = 2*self.precision_score()*self.recall_score() / (self.precision_score()+self.recall_score()+1e-5)
            self.macro_f1 = torch.mean(self.f1)
        return self.f1, self.macro_f1


    def cohen_kappa_score(self):
        if not hasattr(self, 'k'):
            p0 = self.accuracy_score()
            pe = torch.sum(torch.sum(self.confusion_matrix(), dim=0)*torch.sum(self.confusion_matrix(), dim=1))/(self.true.shape[0]**2)
            self.k = (p0-pe)/(1-pe)
        return self.k


    def print_metrics(self):
        cm = self.confusion_matrix().cpu().int().detach().numpy()
        print('confusion_matrix:')
        print(cm)
        print('precision:')
        print(np.round(self.precision_score().cpu().detach().numpy()*100, 2))
        print('recall:')
        print(np.round(self.recall_score().cpu().detach().numpy()*100, 2))
        f1, mf1 = self.f1_score()
        print('f1:')
        print(np.round(f1.cpu().detach().numpy()*100, 2))
        print('acc:')
        print(np.round(self.accuracy_score().item()*100, 2))
        print('mf1:')
        print(np.round(mf1.item()*100, 2))
        print('k:')
        print(np.round(self.cohen_kappa_score().item(), 3))


    def save_metrics(self, save_path, name=None):
        if name != None:
            save_path = os.path.join(save_path, name+'_metrics.xlsx')
        else:
            save_path = os.path.join(save_path, 'metrics.xlsx')
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.merge_cells('A1:A2')
        ws.merge_cells('B1:F1')
        ws.merge_cells('G1:J1')
        ws.merge_cells('J3:J7')
        ws.cell(1, 2, 'Predictions')
        ws.cell(1, 7, 'Metrics')
        for i, s in enumerate(['W', 'N1', 'N2', 'N3', 'R', 'PR', 'RE', 'F1', 'Overall']):
            ws.cell(2, i+2, s)
        for i, s in enumerate(['W', 'N1', 'N2', 'N3', 'R']):
            ws.cell(i+3, 1, s)
        for i in range(Parameter.n_class):
            for j in range(Parameter.n_class):
                ws.cell(i+3, j+2, int(self.confusion_matrix()[i, j].item()))
        for i, pr in enumerate(self.precision_score()):
            ws.cell(i+3, 7, np.round(pr.item()*100, 2))
        for i, re in enumerate(self.recall_score()):
            ws.cell(i+3, 8, np.round(re.item()*100, 2))
        f1, mf1 = self.f1_score()
        for i, _f1 in enumerate(f1):
            ws.cell(i+3, 9, np.round(_f1.item()*100, 2))
        acc = np.round(self.accuracy_score().item()*100, 2)
        mf1 = np.round(mf1.item()*100, 2)
        kappa = np.round(self.cohen_kappa_score().item(), 3)
        val = f'ACC: {acc}\r\nMF1: {mf1}\r\nKappa: {kappa}\r\nTime: '
        ws.cell(3, 10, val)
        wb.save(save_path)